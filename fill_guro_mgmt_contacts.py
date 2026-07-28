# -*- coding: utf-8 -*-
"""건물명_보완검증 시트에 관리사무소·연락처 열 추가 후 네이버/구글 검색으로 채움."""

from __future__ import annotations

import json
import os
import re
import sys
import time
from urllib.parse import quote

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "guro_permits", "구로구_사용승인현황_통합.xlsx")
SHEET = "건물명_보완검증"
CACHE_PATH = os.path.join(BASE, "guro_permits", "mgmt_contact_cache.json")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
}
SESSION = requests.Session()
SESSION.headers.update(HEADERS)

PHONE_RE = re.compile(r"(0\d{1,2}[-.\s]?\d{3,4}[-.\s]?\d{4})")
REJECT = (
    "부동산",
    "공인중개",
    "중개",
    "분양",
    "매매",
    "임대문의",
    "중개사",
    "인테리어",
    "이사",
    "청소업체",
)


def normalize_phone(phone: str) -> str:
    digits = re.sub(r"\D", "", phone)
    if digits.startswith("02"):
        if len(digits) == 9:
            return f"02-{digits[2:5]}-{digits[5:]}"
        if len(digits) == 10:
            return f"02-{digits[2:6]}-{digits[6:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    return phone.strip()


def ok_phone(phone: str) -> bool:
    digits = re.sub(r"\D", "", phone)
    if len(digits) < 9 or len(digits) > 11:
        return False
    if digits.startswith("010") or digits.startswith("000"):
        return False
    return digits.startswith("0")


def is_blank(value: object) -> bool:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return True
    text = str(value).strip()
    return text == "" or text.lower() in {"nan", "미확인", "none"}


def load_cache() -> dict:
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH, encoding="utf-8") as file:
            return json.load(file)
    return {}


def save_cache(cache: dict) -> None:
    with open(CACHE_PATH, "w", encoding="utf-8") as file:
        json.dump(cache, file, ensure_ascii=False, indent=2)


def search_naver(query: str) -> list[dict]:
    response = SESSION.get(
        "https://search.naver.com/search.naver",
        params={"query": query},
        timeout=25,
    )
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    candidates: list[dict] = []

    for box in soup.select(".place_section, .api_subject_bx, .LTgSF, .E2AOU, .cXIkN"):
        text = box.get_text(" ", strip=True)
        if any(token in text for token in REJECT):
            continue
        if "관리" not in text and "전화" not in text:
            continue
        phones = [normalize_phone(p) for p in PHONE_RE.findall(text) if ok_phone(p)]
        title_el = box.select_one("a.place_bluelink, .YwYLL, .TYaxT, .OXiLu, a")
        title = title_el.get_text(" ", strip=True) if title_el else ""
        for phone in phones[:2]:
            candidates.append(
                {
                    "name": title,
                    "phone": phone,
                    "source": "naver",
                    "snippet": text[:180],
                }
            )

    for match in re.finditer(r'"telephone"\s*:\s*"([^"]+)"', html):
        phone = normalize_phone(match.group(1))
        if not ok_phone(phone):
            continue
        chunk = html[max(0, match.start() - 350) : match.end() + 40]
        name_match = re.search(r'"name"\s*:\s*"([^"]{2,80})"', chunk)
        name = name_match.group(1) if name_match else query
        if any(token in name for token in REJECT):
            continue
        candidates.append(
            {"name": name, "phone": phone, "source": "naver_json", "snippet": ""}
        )
    return candidates


def search_google(query: str) -> list[dict]:
    candidates: list[dict] = []
    try:
        response = SESSION.get(
            "https://www.google.com/search",
            params={"q": query, "hl": "ko", "num": 10},
            headers={**HEADERS, "Accept": "text/html"},
            timeout=25,
        )
        if response.status_code != 200 or "unusual traffic" in response.text.lower():
            return candidates
        soup = BeautifulSoup(response.text, "html.parser")
        for block in soup.select("div.g, div.MjjYud, div.N54PNb"):
            text = block.get_text(" ", strip=True)
            if any(token in text for token in REJECT):
                continue
            if "관리" not in text and "전화" not in text:
                continue
            phones = [normalize_phone(p) for p in PHONE_RE.findall(text) if ok_phone(p)]
            title_el = block.select_one("h3")
            title = title_el.get_text(" ", strip=True) if title_el else ""
            for phone in phones[:1]:
                candidates.append(
                    {
                        "name": title,
                        "phone": phone,
                        "source": "google",
                        "snippet": text[:180],
                    }
                )
    except Exception:
        pass
    return candidates


def search_bing(query: str) -> list[dict]:
    """Google 차단 시 보조 검색."""
    candidates: list[dict] = []
    try:
        response = SESSION.get(
            "https://www.bing.com/search",
            params={"q": query, "setlang": "ko"},
            timeout=25,
        )
        soup = BeautifulSoup(response.text, "html.parser")
        for item in soup.select("li.b_algo"):
            text = item.get_text(" ", strip=True)
            if any(token in text for token in REJECT):
                continue
            if "관리" not in text and "전화" not in text:
                continue
            phones = [normalize_phone(p) for p in PHONE_RE.findall(text) if ok_phone(p)]
            title_el = item.select_one("h2")
            title = title_el.get_text(" ", strip=True) if title_el else ""
            for phone in phones[:1]:
                candidates.append(
                    {
                        "name": title,
                        "phone": phone,
                        "source": "bing(google-fallback)",
                        "snippet": text[:180],
                    }
                )
    except Exception:
        pass
    return candidates


def pick(building: str, address: str, candidates: list[dict]) -> dict:
    building_key = re.sub(r"\s+", "", building).lower()
    scored: list[tuple[int, dict]] = []
    for candidate in candidates:
        name = candidate.get("name") or ""
        snippet = candidate.get("snippet") or ""
        blob = (name + " " + snippet).replace(" ", "").lower()
        if any(token in name or token in snippet for token in REJECT):
            continue
        score = 0
        if "관리사무소" in name or "관리사무소" in snippet:
            score += 6
        if "관리실" in name or "관리단" in name or "관리센터" in name:
            score += 4
        if building_key and len(building_key) >= 2 and building_key[:4] in blob:
            score += 4
        if "구로" in blob or "신도림" in blob or "개봉" in blob or "오류" in blob:
            score += 2
        if "구로구" in address and "구로" in blob:
            score += 1
        if candidate.get("source", "").startswith("naver"):
            score += 2
        if candidate.get("source") == "google":
            score += 2
        scored.append((score, candidate))

    scored.sort(key=lambda item: item[0], reverse=True)
    if not scored or scored[0][0] < 5:
        return {
            "관리사무소": "미확인",
            "연락처": "미확인",
            "검색근거": "미확인",
            "신뢰": "low",
        }

    best = scored[0][1]
    office = best.get("name") or ""
    if "관리" not in office:
        office = f"{building} 관리사무소" if building else "관리사무소"
    return {
        "관리사무소": office[:80],
        "연락처": best["phone"],
        "검색근거": f"{best.get('source')}:score{scored[0][0]}",
        "신뢰": "high" if scored[0][0] >= 10 else "medium",
    }


def lookup(building: str, address: str, cache: dict) -> dict:
    key = f"{building}|{address}"
    cached = cache.get(key)
    if cached and not is_blank(cached.get("연락처")):
        return cached

    candidates: list[dict] = []
    queries = []
    if building:
        queries.extend(
            [
                f"{building} 관리사무소",
                f"{building} 구로구 관리사무소 전화번호",
                f"{building} 관리실 전화",
            ]
        )
    if address:
        queries.append(f"{address} 관리사무소")

    for query in queries:
        try:
            candidates.extend(search_naver(query))
        except Exception as exc:
            print(f"  naver err: {exc}", flush=True)
        time.sleep(0.45)
        try:
            google_hits = search_google(query)
            if google_hits:
                candidates.extend(google_hits)
            else:
                candidates.extend(search_bing(query))
        except Exception as exc:
            print(f"  google/bing err: {exc}", flush=True)
        time.sleep(0.35)

    info = pick(building or address, address, candidates)
    info.update({"건물명": building, "대지위치": address})
    cache[key] = info
    return info


def write_sheet(df: pd.DataFrame) -> None:
    book = load_workbook(XLSX)
    if SHEET in book.sheetnames:
        del book[SHEET]
    book.save(XLSX)
    book.close()
    with pd.ExcelWriter(XLSX, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, index=False, sheet_name=SHEET)


def main() -> int:
    df = pd.read_excel(XLSX, sheet_name=SHEET)
    if "관리사무소" not in df.columns:
        df["관리사무소"] = ""
    if "연락처" not in df.columns:
        df["연락처"] = ""
    if "검색근거" not in df.columns:
        df["검색근거"] = ""

    cache = load_cache()
    # unique targets needing fill
    need_mask = df["관리사무소"].map(is_blank) | df["연락처"].map(is_blank)
    targets = df.loc[need_mask].copy()
    print(f"[TARGET] rows needing fill: {int(need_mask.sum())}/{len(df)}", flush=True)

    unique_keys: list[tuple[str, str]] = []
    seen: set[str] = set()
    for _, row in targets.iterrows():
        building = "" if is_blank(row.get("건물명(보완후)")) else str(row.get("건물명(보완후)")).strip()
        address = "" if is_blank(row.get("대지위치")) else str(row.get("대지위치")).strip()
        key = f"{building}|{address}"
        if key in seen:
            continue
        seen.add(key)
        unique_keys.append((building, address))

    print(f"[LOOKUP] unique: {len(unique_keys)}", flush=True)
    for index, (building, address) in enumerate(unique_keys, 1):
        label = building or address
        print(f"[{index}/{len(unique_keys)}] {label}", flush=True)
        info = lookup(building, address, cache)
        print(
            f"  -> {info.get('관리사무소')} | {info.get('연락처')} | {info.get('검색근거')}",
            flush=True,
        )
        if index % 8 == 0:
            save_cache(cache)

    save_cache(cache)

    for idx, row in df.iterrows():
        if not (is_blank(row.get("관리사무소")) or is_blank(row.get("연락처"))):
            continue
        building = "" if is_blank(row.get("건물명(보완후)")) else str(row.get("건물명(보완후)")).strip()
        address = "" if is_blank(row.get("대지위치")) else str(row.get("대지위치")).strip()
        info = cache.get(f"{building}|{address}", {})
        df.at[idx, "관리사무소"] = info.get("관리사무소", "미확인")
        df.at[idx, "연락처"] = info.get("연락처", "미확인")
        df.at[idx, "검색근거"] = info.get("검색근거", "")

    # column order: keep existing then new cols near end
    preferred = [
        "원본인덱스",
        "연면적(㎡)",
        "대지위치",
        "원본건물명",
        "조회건물명",
        "건물명(보완후)",
        "관리사무소",
        "연락처",
        "도로명주소(조회)",
        "판정",
        "검색근거",
        "비고",
    ]
    cols = [c for c in preferred if c in df.columns] + [
        c for c in df.columns if c not in preferred
    ]
    df = df[cols]
    write_sheet(df)

    filled = int((~df["연락처"].map(is_blank)).sum())
    print(f"[DONE] filled contacts={filled}/{len(df)} -> {XLSX}", flush=True)
    print(df["연락처"].map(lambda x: "미확인" if is_blank(x) else "채움").value_counts().to_string())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
