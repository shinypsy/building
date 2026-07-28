# -*- coding: utf-8 -*-
"""연면적 5,000~30,000㎡ 필터 → 새 시트 내림차순."""
import os
import sys
import re

import pandas as pd
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
# 실제 폴더명(geumcheon) 우선, 과거 yangcheon 폴더도 허용
_CANDIDATES = [
    os.path.join(BASE, "geumcheon", "금천구_사용승인현황_2016-2026.xlsx"),
    os.path.join(BASE, "yangcheon", "금천구_사용승인현황_2016-2026.xlsx"),
]
PATH = next((p for p in _CANDIDATES if os.path.exists(p)), _CANDIDATES[0])
SRC_SHEET = "사용승인_통합"
NEW_SHEET = "중형_연면적5천~3만"


def to_area(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace(" ", "")
    s = s.replace("㎡", "").replace("m2", "").replace("M2", "")
    if not s or s.lower() in ("nan", "none", "-"):
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def main() -> int:
    if not os.path.exists(PATH):
        print(f"파일 없음: {PATH}")
        return 1

    df = pd.read_excel(PATH, sheet_name=SRC_SHEET)
    area_cols = [c for c in df.columns if "연면적" in str(c) and "증축" not in str(c)]
    if not area_cols:
        area_cols = [c for c in df.columns if "연면적" in str(c)]
    if not area_cols:
        print("연면적 컬럼을 찾지 못했습니다:", list(df.columns)[:30])
        return 1

    area_col = area_cols[0]
    print(f"연면적 컬럼: {area_col}")

    areas = df[area_col].map(to_area)
    mask = areas.map(lambda x: x is not None and 5000 <= x <= 30000)
    filtered = df.loc[mask].copy()
    filtered["_연면적_수치"] = areas[mask]
    filtered = filtered.sort_values("_연면적_수치", ascending=False)
    filtered = filtered.drop(columns=["_연면적_수치"])
    filtered = filtered.reset_index(drop=True)

    print(f"원본: {len(df)}행 → 필터: {len(filtered)}행")
    if len(filtered):
        sample = filtered[area_col].head(5).tolist()
        print("상위 연면적 샘플:", sample)

    # 기존 통합 시트 유지 + 새 시트 추가/교체
    with pd.ExcelWriter(
        PATH,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        filtered.to_excel(writer, sheet_name=NEW_SHEET, index=False)

    # 시트 존재 확인
    wb = load_workbook(PATH, read_only=True)
    print("시트:", wb.sheetnames)
    wb.close()
    print(f"완료: {PATH} / '{NEW_SHEET}'")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
