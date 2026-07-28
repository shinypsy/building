# -*- coding: utf-8 -*-
"""구로구 사용승인 건물명 보완·검증 (연면적 5,000㎡ 이상, map.naver.com)."""

from __future__ import annotations

import json
import os
import re
import sys
import time
from urllib.parse import quote

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "guro_permits", "구로구_사용승인현황_통합.xlsx")
CACHE_PATH = os.path.join(BASE, "guro_permits", "building_name_cache.json")
AREA_MIN = 5000.0
SOURCE_SHEET = "사용승인_통합"
RESULT_SHEET = "건물명_보완검증"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
}

REJECT_NAMES = (
    "네이버",
    "지도",
    "검색",
    "더보기",
    "광고",
    "블로그",
    "카페",
    "뉴스",
    "부동산",
    "공인중개",
    "나무위키",
    "위키",
    "ATM",
    "쏘카",
)


def is_blank_name(value: object) -> bool:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return True
    text = str(value).strip()
    return text == "" or text.lower() == "nan"


def normalize_name(value: object) -> str:
    if is_blank_name(value):
        return ""
    text = str(value)
    text = re.sub(r"\(.*?\)", " ", text)
    text = re.sub(r"[^0-9A-Za-z가-힣]", "", text)
    return text.lower()


def clean_candidate(name: str) -> str:
    name = re.sub(r"\s+", " ", name).strip()
    if len(name) < 2 or len(name) > 60:
        return ""
    if any(token in name for token in REJECT_NAMES):
        return ""
    if name.startswith("서울") and ("구로구" in name or "특별시" in name):
        return ""
    return name


def load_cache() -> dict:
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH, encoding="utf-8") as file:
            return json.load(file)
    return {}


def save_cache(cache: dict) -> None:
    with open(CACHE_PATH, "w", encoding="utf-8") as file:
        json.dump(cache, file, ensure_ascii=False, indent=2)


def pick_best_name(candidates: list[str], address: str) -> str:
    addr_norm = normalize_name(address)
    scored: list[tuple[int, str]] = []
    for raw in candidates:
        name = clean_candidate(raw)
        if not name:
            continue
        score = 0
        norm = normalize_name(name)
        if "아파트" in name or "타워" in name or "빌딩" in name or "센터" in name:
            score += 2
        if "시티" in name or "타운" in name or "플라자" in name:
            score += 1
        if norm and norm not in addr_norm:
            score += 3
        if 2 <= len(name) <= 25:
            score += 2
        if re.search(r"[가-힣A-Za-z]", name):
            score += 1
        # map.naver build_name is usually a clean short label
        if re.fullmatch(r"[가-힣A-Za-z0-9 ·&\-]{2,30}", name):
            score += 2
        scored.append((score, name))
    if not scored:
        return ""
    scored.sort(key=lambda item: (-item[0], len(item[1])))
    return scored[0][1]


def lookup_map_naver(page, address: str) -> dict:
    url = f"https://map.naver.com/p/search/{quote(address)}"
    result = {"name": "", "road": "", "url": url, "source": "map.naver", "note": ""}
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=45000)
        page.wait_for_timeout(4500)
        names: list[str] = []
        roads: list[str] = []
        for frame in page.frames:
            try:
                html = frame.content()
            except Exception:
                continue
            for sel in [
                ".build_name",
                "span.TYaxT",
                "a.place_bluelink",
                ".YwYLL",
                ".place_bluelink",
                "strong.SEARCH_TITLE",
                ".search_title",
                ".title_area",
                ".Fc1rA",
            ]:
                try:
                    for el in frame.query_selector_all(sel):
                        text = el.inner_text().strip()
                        if text:
                            names.append(text)
                except Exception:
                    continue
            for sel in [
                ".label_address_road + span",
                ".LDgIH",
                ".jO09N",
                ".Pb4bU",
                ".addr",
            ]:
                try:
                    for el in frame.query_selector_all(sel):
                        text = el.inner_text().strip()
                        if text:
                            roads.append(text)
                except Exception:
                    continue
            # Prefer official map address-panel building name
            build_names = re.findall(r'class="build_name">([^<]{2,80})<', html)
            if build_names:
                preferred = clean_candidate(build_names[0])
                if preferred:
                    names.insert(0, preferred)
            names.extend(build_names[:10])
            roads.extend(
                re.findall(
                    r'label_address_road.*?</span><span>([^<]{2,80})</span>',
                    html,
                    flags=re.S,
                )[:5]
            )
            names.extend(re.findall(r'"name"\s*:\s*"([^"]{2,80})"', html)[:20])
            roads.extend(re.findall(r'"roadAddress"\s*:\s*"([^"]{5,120})"', html)[:10])
            if "ncaptcha-iframe" in html.lower():
                # address panel may still render with captcha frame present
                if not result["note"]:
                    result["note"] = "captcha_frame"
        result["name"] = pick_best_name(names, address)
        result["road"] = roads[0] if roads else ""
        if not result["name"]:
            result["note"] = (result["note"] + ";empty").strip(";")
    except Exception as exc:
        result["note"] = f"map_error:{exc}"
    return result


def lookup_search_naver(address: str) -> dict:
    url = f"https://search.naver.com/search.naver?query={quote(address)}"
    result = {"name": "", "road": "", "url": url, "source": "search.naver", "note": "fallback"}
    try:
        response = requests.get(
            "https://search.naver.com/search.naver",
            params={"query": address},
            headers=HEADERS,
            timeout=25,
        )
        response.raise_for_status()
        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        names: list[str] = []
        for sel in [
            ".place_bluelink",
            ".TYaxT",
            ".OXiLu",
            ".sds-comps-text-type-headline1",
            ".total_tit",
        ]:
            for el in soup.select(sel):
                text = el.get_text(" ", strip=True)
                if text:
                    names.append(text)
        names.extend(re.findall(r'"name"\s*:\s*"([^"]{2,60})"', html)[:30])
        # Prefer patterns like "주소 건물명"
        for match in re.finditer(
            re.escape(address.split("구로구")[-1].strip()) + r"\s*([가-힣A-Za-z0-9·\-]{2,40})",
            html,
        ):
            names.append(match.group(1))
        result["name"] = pick_best_name(names, address)
        roads = re.findall(r"(서울[^\s\"<>]{5,40})", html)
        result["road"] = roads[0] if roads else ""
        if not result["name"]:
            result["note"] = "fallback;empty"
    except Exception as exc:
        result["note"] = f"search_error:{exc}"
    return result


def judge(original: str, found: str) -> str:
    if not found:
        return "확인불가"
    if is_blank_name(original):
        return "채움"
    left = normalize_name(original)
    right = normalize_name(found)
    if not left or not right:
        return "확인불가"
    # soft normalize: drop common org words
    for token in ("부속", "주식회사", "유한회사", "학교법인"):
        left = left.replace(token, "")
        right = right.replace(token, "")
    if left in right or right in left:
        return "일치"
    left_tokens = set(re.findall(r"[가-힣A-Za-z0-9]{2,}", str(original)))
    right_tokens = set(re.findall(r"[가-힣A-Za-z0-9]{2,}", found))
    stop = {"서울", "특별시", "구로구", "구로동", "건물", "업무시설", "판매시설"}
    left_tokens -= stop
    right_tokens -= stop
    if left_tokens and right_tokens and left_tokens & right_tokens:
        return "일치"
    return "불일치"


def lookup_address(page, address: str, cache: dict) -> dict:
    if address in cache:
        cached = dict(cache[address])
        cached["note"] = ((cached.get("note") or "") + ";cache").strip(";")
        return cached

    mapped = lookup_map_naver(page, address)
    if mapped.get("name"):
        cache[address] = mapped
        return mapped

    searched = lookup_search_naver(address)
    if mapped.get("note"):
        searched["note"] = f"{mapped['note']};{searched.get('note','')}".strip(";")
    searched["map_url"] = mapped.get("url", "")
    cache[address] = searched
    return searched


def write_result_sheet(df_result: pd.DataFrame) -> None:
    book = load_workbook(XLSX)
    if RESULT_SHEET in book.sheetnames:
        del book[RESULT_SHEET]
    book.save(XLSX)
    book.close()

    with pd.ExcelWriter(XLSX, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_result.to_excel(writer, index=False, sheet_name=RESULT_SHEET)


def main() -> int:
    df = pd.read_excel(XLSX, sheet_name=SOURCE_SHEET)
    area = pd.to_numeric(df["연면적(㎡)"], errors="coerce")
    target = df[area >= AREA_MIN].copy()
    target["_area"] = area[area >= AREA_MIN]
    target = target.sort_values("_area", ascending=False)

    print(f"[TARGET] 연면적>={AREA_MIN}: {len(target)}행", flush=True)
    cache = load_cache()
    rows: list[dict] = []

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        unique_addrs = (
            target["대지위치"].astype(str).str.strip().replace({"nan": ""}).tolist()
        )
        seen: set[str] = set()
        lookup_plan: list[str] = []
        for address in unique_addrs:
            if not address or address in seen:
                continue
            seen.add(address)
            lookup_plan.append(address)

        print(f"[LOOKUP] 고유주소 {len(lookup_plan)}건", flush=True)
        for index, address in enumerate(lookup_plan, 1):
            print(f"[{index}/{len(lookup_plan)}] {address}", flush=True)
            result = lookup_address(page, address, cache)
            print(
                f"  -> {result.get('name') or '(없음)'} "
                f"[{result.get('source')}] {result.get('note')}",
                flush=True,
            )
            if index % 10 == 0:
                save_cache(cache)
            time.sleep(0.8)
        browser.close()

    save_cache(cache)

    for idx, row in target.iterrows():
        address = str(row.get("대지위치", "")).strip()
        original = "" if is_blank_name(row.get("건물명")) else str(row.get("건물명")).strip()
        found_data = cache.get(address, {})
        found_name = found_data.get("name", "")
        verdict = judge(original, found_name)
        filled_name = found_name if verdict == "채움" else original
        rows.append(
            {
                "원본인덱스": int(idx) + 2,
                "연면적(㎡)": row.get("연면적(㎡)"),
                "대지위치": address,
                "원본건물명": original,
                "조회건물명": found_name,
                "건물명(보완후)": filled_name if verdict == "채움" else (original or found_name),
                "도로명주소(조회)": found_data.get("road", ""),
                "판정": verdict,
                "근거URL": found_data.get("url") or found_data.get("map_url", ""),
                "조회출처": found_data.get("source", ""),
                "비고": found_data.get("note", ""),
            }
        )

    result_df = pd.DataFrame(rows)
    write_result_sheet(result_df)

    counts = result_df["판정"].value_counts().to_dict()
    print(f"\n[DONE] sheet={RESULT_SHEET}")
    print(f"  rows={len(result_df)}")
    print(f"  판정={counts}")
    print(f"  file={XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
